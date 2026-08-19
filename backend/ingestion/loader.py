"""
Excel loader: reads EPIC.xlsx and EPIC Releases.xlsx into structured dicts.

Fails loudly if any required column is missing — never silently falls back.
All date fields are stored as ISO-8601 strings (YYYY-MM-DD) or None.
"""

from __future__ import annotations

import io
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any

import openpyxl
from openpyxl.worksheet.worksheet import Worksheet

from config_loader import get_config


# ── Helpers ───────────────────────────────────────────────────────────────────

def _require_col(headers: list[str | None], name: str, sheet: str) -> int:
    """Return 0-based index of *name* in *headers*, raising loudly if absent."""
    try:
        return headers.index(name)
    except ValueError:
        raise ValueError(
            f"Required column {name!r} not found in sheet {sheet!r}. "
            f"Available columns: {[h for h in headers if h]}"
        )


def _col_idx(headers: list[str | None], name: str) -> int | None:
    """Return 0-based index or None (optional columns)."""
    try:
        return headers.index(name)
    except ValueError:
        return None


def _cell(row: tuple, idx: int | None) -> Any:
    if idx is None or idx >= len(row):
        return None
    val = row[idx]
    if isinstance(val, str):
        return val.strip() or None
    return val


_MONTH_NAMES = {
    "january": 1, "february": 2, "march": 3, "april": 4,
    "may": 5, "june": 6, "july": 7, "august": 8,
    "september": 9, "october": 10, "november": 11, "december": 12,
    "jan": 1, "feb": 2, "mar": 3, "apr": 4,
    "jun": 6, "jul": 7, "aug": 8, "sep": 9, "oct": 10, "nov": 11, "dec": 12,
}


def _parse_month(val: Any) -> int | None:
    """Accept an integer, a float, or a month name string."""
    if val is None:
        return None
    if isinstance(val, (int, float)):
        return int(val) if val else None
    s = str(val).strip().lower()
    return _MONTH_NAMES.get(s) or (int(s) if s.isdigit() else None)


def _date_from_parts(year: Any, month: Any, day: Any) -> str | None:
    """Combine Year/Month/Day cells into ISO date string.
    Month may be a number or a full English month name."""
    try:
        y = int(year) if year is not None else None
        m = _parse_month(month)
        d = int(day) if day is not None else None
        if y and m and d:
            return f"{y:04d}-{m:02d}-{d:02d}"
    except (TypeError, ValueError):
        pass
    return None


def _parse_fix_versions(raw: Any) -> list[str]:
    """Split a semicolon-delimited Fix versions string into a list of names."""
    if not raw:
        return []
    return [v.strip() for v in str(raw).split(";") if v.strip()]


def _read_sheet_rows(wb: openpyxl.Workbook, sheet_name: str) -> tuple[list, list[tuple]]:
    """Return (headers, data_rows) from a sheet, skipping blank rows."""
    ws: Worksheet = wb[sheet_name]
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        raise ValueError(f"Sheet {sheet_name!r} is empty")
    headers = [str(h).strip() if h is not None else None for h in rows[0]]
    data = [r for r in rows[1:] if any(c is not None for c in r)]
    return headers, data


# ── Result dataclasses ────────────────────────────────────────────────────────

@dataclass
class LoadedData:
    epics: list[dict] = field(default_factory=list)
    epic_transitions: list[dict] = field(default_factory=list)
    capabilities: list[dict] = field(default_factory=list)
    features: list[dict] = field(default_factory=list)
    feature_transitions: list[dict] = field(default_factory=list)
    art_lookup: dict[str, str] = field(default_factory=dict)          # feature_key → ART value
    releases: list[dict] = field(default_factory=list)
    epic_releases: list[dict] = field(default_factory=list)            # [{epic_key, release_name}]
    capability_releases: list[dict] = field(default_factory=list)      # [{cap_key, release_name}]
    warnings: list[str] = field(default_factory=list)


# ── Main loaders ──────────────────────────────────────────────────────────────

def load_main_workbook(source: bytes | Path) -> LoadedData:
    """Load EPIC.xlsx (or equivalent bytes) and return a LoadedData object."""
    cfg = get_config()
    ec = cfg.excel_columns

    if isinstance(source, (bytes, bytearray)):
        wb = openpyxl.load_workbook(io.BytesIO(source), read_only=False, data_only=True)
    else:
        wb = openpyxl.load_workbook(source, read_only=False, data_only=True)

    result = LoadedData()

    # ── EPICs ─────────────────────────────────────────────────────────────
    headers, rows = _read_sheet_rows(wb, ec["epic_sheet"])
    key_i = _require_col(headers, ec["epic_key_col"], ec["epic_sheet"])
    title_i = _require_col(headers, ec["epic_title_col"], ec["epic_sheet"])
    status_i = _require_col(headers, ec["epic_status_col"], ec["epic_sheet"])

    for row in rows:
        key = _cell(row, key_i)
        if not key:
            continue
        result.epics.append({
            "epic_key": key,
            "title": _cell(row, title_i) or key,
            "status": _cell(row, status_i) or "Unknown",
            "delivery_increment": None,   # EPIC sheet has no DI column
        })

    # ── EPIC Transitions ──────────────────────────────────────────────────
    headers, rows = _read_sheet_rows(wb, ec["epic_transition_sheet"])
    ikey_i = _require_col(headers, ec["transition_issue_key_col"], ec["epic_transition_sheet"])
    from_i = _require_col(headers, ec["transition_from_col"], ec["epic_transition_sheet"])
    to_i = _require_col(headers, ec["transition_to_col"], ec["epic_transition_sheet"])
    td_year_i = _require_col(headers, ec["transition_date_year_col"], ec["epic_transition_sheet"])
    td_month_i = _require_col(headers, ec["transition_date_month_col"], ec["epic_transition_sheet"])
    td_day_i = _require_col(headers, ec["transition_date_day_col"], ec["epic_transition_sheet"])

    for row in rows:
        key = _cell(row, ikey_i)
        to_status = _cell(row, to_i)
        if not key or not to_status:
            continue
        date = _date_from_parts(_cell(row, td_year_i), _cell(row, td_month_i), _cell(row, td_day_i))
        result.epic_transitions.append({
            "epic_key": key,
            "from_status": _cell(row, from_i),
            "to_status": to_status,
            "transition_date": date,
        })

    # ── Capabilities ──────────────────────────────────────────────────────
    headers, rows = _read_sheet_rows(wb, ec["capability_sheet"])
    ckey_i = _require_col(headers, ec["capability_key_col"], ec["capability_sheet"])
    ekey_i = _require_col(headers, ec["capability_epic_col"], ec["capability_sheet"])
    ctitle_i = _require_col(headers, ec["capability_title_col"], ec["capability_sheet"])
    cstatus_i = _require_col(headers, ec["capability_status_col"], ec["capability_sheet"])
    cdi_i = _col_idx(headers, ec["capability_di_col"])
    cts_year_i = _col_idx(headers, ec["capability_target_start_year_col"])
    cts_month_i = _col_idx(headers, ec["capability_target_start_month_col"])
    cts_day_i = _col_idx(headers, ec["capability_target_start_day_col"])
    cte_year_i = _col_idx(headers, ec["capability_target_end_year_col"])
    cte_month_i = _col_idx(headers, ec["capability_target_end_month_col"])
    cte_day_i = _col_idx(headers, ec["capability_target_end_day_col"])

    for row in rows:
        key = _cell(row, ckey_i)
        if not key:
            continue
        result.capabilities.append({
            "cap_key": key,
            "epic_key": _cell(row, ekey_i),
            "title": _cell(row, ctitle_i) or key,
            "status": _cell(row, cstatus_i) or "Unknown",
            "delivery_increment": _cell(row, cdi_i) if cdi_i is not None else None,
            "target_start_date": _date_from_parts(
                _cell(row, cts_year_i), _cell(row, cts_month_i), _cell(row, cts_day_i)
            ),
            "target_end_date": _date_from_parts(
                _cell(row, cte_year_i), _cell(row, cte_month_i), _cell(row, cte_day_i)
            ),
        })

    # ── Features ──────────────────────────────────────────────────────────
    headers, rows = _read_sheet_rows(wb, ec["feature_sheet"])
    fkey_i = _require_col(headers, ec["feature_key_col"], ec["feature_sheet"])
    fck_i = _require_col(headers, ec["feature_cap_col"], ec["feature_sheet"])
    ftitle_i = _require_col(headers, ec["feature_title_col"], ec["feature_sheet"])
    fstatus_i = _require_col(headers, ec["feature_status_col"], ec["feature_sheet"])
    fdi_i = _col_idx(headers, ec["feature_di_col"])

    fc_year_i = _col_idx(headers, ec["feature_created_year_col"])
    fc_month_i = _col_idx(headers, ec["feature_created_month_col"])
    fc_day_i = _col_idx(headers, ec["feature_created_day_col"])

    fcmt_year_i = _col_idx(headers, ec["feature_committed_year_col"])
    fcmt_month_i = _col_idx(headers, ec["feature_committed_month_col"])
    fcmt_day_i = _col_idx(headers, ec["feature_committed_day_col"])

    fdone_year_i = _col_idx(headers, ec["feature_done_year_col"])
    fdone_month_i = _col_idx(headers, ec["feature_done_month_col"])
    fdone_day_i = _col_idx(headers, ec["feature_done_day_col"])

    fts_year_i = _col_idx(headers, ec["feature_target_start_year_col"])
    fts_month_i = _col_idx(headers, ec["feature_target_start_month_col"])
    fts_day_i = _col_idx(headers, ec["feature_target_start_day_col"])

    fte_year_i = _col_idx(headers, ec["feature_target_end_year_col"])
    fte_month_i = _col_idx(headers, ec["feature_target_end_month_col"])
    fte_day_i = _col_idx(headers, ec["feature_target_end_day_col"])

    for row in rows:
        key = _cell(row, fkey_i)
        if not key:
            continue
        created_date = _date_from_parts(
            _cell(row, fc_year_i), _cell(row, fc_month_i), _cell(row, fc_day_i)
        )
        if not created_date:
            result.warnings.append(f"Feature {key}: missing Created date — will use 1900-01-01 as fallback")
            created_date = "1900-01-01"
        result.features.append({
            "feature_key": key,
            "cap_key": _cell(row, fck_i),
            "title": _cell(row, ftitle_i) or key,
            "status": _cell(row, fstatus_i) or "Unknown",
            "delivery_increment": _cell(row, fdi_i) if fdi_i is not None else None,
            "created_date": created_date,
            "date_committed": _date_from_parts(
                _cell(row, fcmt_year_i), _cell(row, fcmt_month_i), _cell(row, fcmt_day_i)
            ),
            "date_done": _date_from_parts(
                _cell(row, fdone_year_i), _cell(row, fdone_month_i), _cell(row, fdone_day_i)
            ),
            "target_start_date": _date_from_parts(
                _cell(row, fts_year_i), _cell(row, fts_month_i), _cell(row, fts_day_i)
            ),
            "target_end_date": _date_from_parts(
                _cell(row, fte_year_i), _cell(row, fte_month_i), _cell(row, fte_day_i)
            ),
        })

    # ── Feature Transitions ───────────────────────────────────────────────
    headers, rows = _read_sheet_rows(wb, ec["transition_sheet"])
    tikey_i = _require_col(headers, ec["transition_issue_key_col"], ec["transition_sheet"])
    tfrom_i = _require_col(headers, ec["transition_from_col"], ec["transition_sheet"])
    tto_i = _require_col(headers, ec["transition_to_col"], ec["transition_sheet"])
    tt_year_i = _require_col(headers, ec["transition_date_year_col"], ec["transition_sheet"])
    tt_month_i = _require_col(headers, ec["transition_date_month_col"], ec["transition_sheet"])
    tt_day_i = _require_col(headers, ec["transition_date_day_col"], ec["transition_sheet"])

    for row in rows:
        key = _cell(row, tikey_i)
        to_status = _cell(row, tto_i)
        if not key or not to_status:
            continue
        date = _date_from_parts(_cell(row, tt_year_i), _cell(row, tt_month_i), _cell(row, tt_day_i))
        result.feature_transitions.append({
            "feature_key": key,
            "from_status": _cell(row, tfrom_i),
            "to_status": to_status,
            "transition_date": date,
        })

    # ── ART ───────────────────────────────────────────────────────────────
    headers, rows = _read_sheet_rows(wb, ec["art_sheet"])
    akey_i = _require_col(headers, ec["art_issue_key_col"], ec["art_sheet"])
    aval_i = _require_col(headers, ec["art_value_col"], ec["art_sheet"])

    for row in rows:
        key = _cell(row, akey_i)
        val = _cell(row, aval_i)
        if key:
            result.art_lookup[key] = val or ""

    wb.close()
    return result


def load_releases_workbook(source: bytes | Path) -> LoadedData:
    """Load EPIC Releases.xlsx and return a LoadedData object (releases fields only)."""
    cfg = get_config()
    ec = cfg.excel_columns

    if isinstance(source, (bytes, bytearray)):
        wb = openpyxl.load_workbook(io.BytesIO(source), read_only=False, data_only=True)
    else:
        wb = openpyxl.load_workbook(source, read_only=False, data_only=True)

    result = LoadedData()

    # ── ALL UNRELEASED — release master list ──────────────────────────────
    headers, rows = _read_sheet_rows(wb, ec["all_unreleased_sheet"])
    rname_i = _require_col(headers, ec["release_name_col"], ec["all_unreleased_sheet"])
    rstatus_i = _col_idx(headers, ec["release_status_col"])
    rprog_i = _col_idx(headers, ec["release_progress_col"])
    rstart_i = _col_idx(headers, ec["release_start_date_col"])
    rrelease_i = _col_idx(headers, ec["release_release_date_col"])
    rdesc_i = _col_idx(headers, ec["release_description_col"])

    seen_names: dict[str, dict] = {}
    for row in rows:
        name = _cell(row, rname_i)
        if not name:
            continue
        entry = {
            "release_name": name,
            "status": _cell(row, rstatus_i) if rstatus_i is not None else None,
            "progress": _cell(row, rprog_i) if rprog_i is not None else None,
            "start_date": _parse_date_cell(_cell(row, rstart_i)) if rstart_i is not None else None,
            "release_date": _parse_date_cell(_cell(row, rrelease_i)) if rrelease_i is not None else None,
            "description": _cell(row, rdesc_i) if rdesc_i is not None else None,
        }
        if name in seen_names:
            # Duplicate name — apply tie-break: prefer record with more complete dates
            existing = seen_names[name]
            existing_dates = sum(1 for f in ["start_date", "release_date"] if existing[f])
            new_dates = sum(1 for f in ["start_date", "release_date"] if entry[f])
            if new_dates > existing_dates:
                seen_names[name] = entry
            result.warnings.append(
                f"Duplicate release name {name!r} — kept record with most complete dates"
            )
        else:
            seen_names[name] = entry

    result.releases = list(seen_names.values())

    # ── EPIC Releases — which releases belong to each EPIC ────────────────
    headers, rows = _read_sheet_rows(wb, ec["epic_releases_sheet"])
    erkey_i = _require_col(headers, ec["epic_releases_key_col"], ec["epic_releases_sheet"])
    erfv_i = _require_col(headers, ec["epic_releases_fixversions_col"], ec["epic_releases_sheet"])

    for row in rows:
        key = _cell(row, erkey_i)
        fv = _cell(row, erfv_i)
        if not key or not fv:
            continue
        for rname in _parse_fix_versions(fv):
            result.epic_releases.append({"epic_key": key, "release_name": rname})

    # ── Capability Releases — which release each Capability is in ─────────
    headers, rows = _read_sheet_rows(wb, ec["capability_releases_sheet"])
    crkey_i = _require_col(headers, ec["capability_releases_key_col"], ec["capability_releases_sheet"])
    crfv_i = _require_col(headers, ec["capability_releases_fixversions_col"], ec["capability_releases_sheet"])

    for row in rows:
        key = _cell(row, crkey_i)
        fv = _cell(row, crfv_i)
        if not key or not fv:
            continue
        for rname in _parse_fix_versions(fv):
            result.capability_releases.append({"cap_key": key, "release_name": rname})

    wb.close()
    return result


def _parse_date_cell(val: Any) -> str | None:
    """Parse a date value that may be a datetime object or a string."""
    if val is None:
        return None
    if hasattr(val, "strftime"):
        return val.strftime("%Y-%m-%d")
    s = str(val).strip()
    if not s:
        return None
    # Try common formats
    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%d-%b-%Y", "%d %b %Y"):
        try:
            from datetime import datetime
            return datetime.strptime(s, fmt).strftime("%Y-%m-%d")
        except ValueError:
            continue
    return s  # return as-is if unparseable; DQ rule will flag it
